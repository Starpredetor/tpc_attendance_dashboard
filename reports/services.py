from io import BytesIO
from urllib.parse import quote

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.utils.text import slugify
from collections import defaultdict

from lectures.models import Lecture
from students.models import Student
from attendance.models import AttendanceRecord

from django.utils import timezone

def make_naive(dt):
    return timezone.make_naive(dt) if timezone.is_aware(dt) else dt

PRESENT_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
ABSENT_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def generate_student_attendance_excel(student):
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    headers = [
        "Lecture Date",
        "Session",
        "Batch",
        "Lecture Title",
        "Status",
        "Marked At",
        "Last Updated",
    ]

    ws.append(headers)

    # Header styling
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

    records = (
        AttendanceRecord.objects
        .filter(student=student)
        .select_related("lecture", "lecture__batch",)
        .order_by("lecture__date", "lecture__lecture_type")
    )

    for record in records:
        lecture = record.lecture
        session_name = lecture.get_lecture_type_display()

        ws.append([
            lecture.date,
            session_name,
            lecture.batch.name,
            lecture.title or "-",
            "Present" if record.status == "P" else "Absent",
            make_naive(record.marked_at,),
            make_naive(record.updated_at) if hasattr(record, "updated_at") else make_naive(record.marked_at),
        ])

        # Color the Status cell (column 5)
        status_cell = ws.cell(row=ws.max_row, column=5)
        status_cell.fill = PRESENT_FILL if record.status == "P" else ABSENT_FILL
        # Borders for the newly added row
        for col in range(1, ws.max_column + 1):
            ws.cell(row=ws.max_row, column=col).border = THIN_BORDER

    ws.freeze_panes = "A2"

    # Number formats and reasonable widths
    date_fmt = "YYYY-MM-DD"
    dt_fmt = "YYYY-MM-DD hh:mm"
    widths = [15, 12, 18, 30, 12, 20, 20]
    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = widths[col - 1] if col - 1 < len(widths) else 18
        if col == 1:
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=col).number_format = date_fmt
        if col in (6, 7):
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=col).number_format = dt_fmt

    # Build response via BytesIO and sanitize filename
    safe_base = slugify(f"{getattr(student, 'roll_number', '')}_{student.full_name}") or "student"
    filename = f"{safe_base}_attendance_report.xlsx"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    # RFC 5987 filename* for UTF-8 support
    response["Content-Disposition"] = (
        f"attachment; filename={filename}; filename*=UTF-8''{quote(filename)}"
    )
    response["X-Content-Type-Options"] = "nosniff"
    return response

def generate_lecture_attendance_matrix_excel(batches, start_date, end_date):
    wb = Workbook()
    wb.remove(wb.active)

    for batch in batches:
        ws = wb.create_sheet(title=batch.name[:31])

        lectures = (
            Lecture.objects
            .filter(batch=batch, date__range=(start_date, end_date))
            .order_by("date", "lecture_type")
        )

        if not lectures.exists():
            continue

        lecture_map = defaultdict(dict)
        for lec in lectures:
            lecture_map[lec.date][lec.lecture_type] = lec

        dates = sorted(lecture_map.keys())

        students = (
            Student.objects
            .filter(batch=batch, is_active=True)
            .order_by("roll_number")
        )

        fixed_headers = ["Roll No", "Name", "Branch", "Attendance %"]

        # ----- HEADER ROWS -----
        ws.append(fixed_headers)  # Row 1
        ws.append([None] * len(fixed_headers))  # Row 2 placeholder for merged headers

        # Merge the first 4 header cells over two rows and style them
        for col in range(1, len(fixed_headers) + 1):
            ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)
            c = ws.cell(row=1, column=col)
            c.value = fixed_headers[col - 1]
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.fill = HEADER_FILL
            c.border = THIN_BORDER

        col = len(fixed_headers) + 1
        for d in dates:
            ws.merge_cells(
                start_row=1,
                start_column=col,
                end_row=1,
                end_column=col + 1,
            )
            ws.cell(row=1, column=col).value = d.strftime("%d-%b-%Y")
            ws.cell(row=2, column=col).value = "Morning"
            ws.cell(row=2, column=col + 1).value = "Afternoon"
            col += 2

        # Style date headers and session headers
        for c in range(len(fixed_headers) + 1, ws.max_column + 1):
            top = ws.cell(row=1, column=c)
            bottom = ws.cell(row=2, column=c)
            if top.value:  # date cell
                top.font = Font(bold=True)
                top.alignment = Alignment(horizontal="center", vertical="center")
                top.fill = HEADER_FILL
                top.border = THIN_BORDER
            bottom.alignment = Alignment(horizontal="center", vertical="center")
            bottom.fill = HEADER_FILL
            bottom.border = THIN_BORDER

        ws.freeze_panes = "E3"

        attendance_lookup = {
            (r.student_id, r.lecture_id): r.status
            for r in AttendanceRecord.objects.filter(
                lecture__batch=batch,
                lecture__date__range=(start_date, end_date)
            )
        }

        row = 3
        for student in students:
            ws.append([student.roll_number, student.full_name, student.branch.name, ""])  # Row added
            # Style fixed columns for the row
            for col in range(1, 5):
                cell = ws.cell(row=row, column=col)
                cell.border = THIN_BORDER
                if col in (1, 4):
                    cell.alignment = Alignment(horizontal="center")
                else:
                    cell.alignment = Alignment(horizontal="left")

            present_count = 0
            total_sessions = 0

            col = len(fixed_headers) + 1
            for d in dates:
                for session in ["MS", "AS"]:
                    lec = lecture_map[d].get(session)
                    cell = ws.cell(row=row, column=col)

                    if lec:
                        status = attendance_lookup.get(
                            (student.id, lec.id), "A"
                        )
                        cell.value = status
                        cell.fill = PRESENT_FILL if status == "P" else ABSENT_FILL
                        cell.alignment = Alignment(horizontal="center")
                        cell.border = THIN_BORDER

                        total_sessions += 1
                        if status == "P":
                            present_count += 1
                    else:
                        cell.value = "-"
                        cell.alignment = Alignment(horizontal="center")
                        cell.border = THIN_BORDER

                    col += 1

            # Use a true percentage value and number format
            percent_cell = ws.cell(row=row, column=4)
            if total_sessions > 0:
                percent_cell.value = present_count / total_sessions
            else:
                percent_cell.value = None
            percent_cell.number_format = "0.00%"
            percent_cell.alignment = Alignment(horizontal="center")
            percent_cell.border = THIN_BORDER

            row += 1

        # Column widths: fixed columns specific, others moderate
        widths = {1: 10, 2: 28, 3: 14, 4: 14}
        for i in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(i)].width = widths.get(i, 12)

        # Autofilter across the sheet
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

    # Prepare response via BytesIO and safe filename
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    safe_batch = slugify("_".join(batch.name for batch in batches)) or "batches"
    safe_range = slugify(f"{start_date}_to_{end_date}") or "range"
    filename = f"{safe_range}_{safe_batch}_attendance_report.xlsx"

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        f"attachment; filename={filename}; filename*=UTF-8''{quote(filename)}"
    )
    response["X-Content-Type-Options"] = "nosniff"
    return response
