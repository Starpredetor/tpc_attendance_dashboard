"""
Utility script to create a sample Excel template for attendance import.
Run: python create_sample_attendance_template.py
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime, timedelta

def create_template():
    wb = Workbook()
    wb.remove(wb.active)
    
    batches = ["Batch 1", "Batch 2", "Batch 3", "Batch 4"]
    
    for batch_name in batches:
        ws = wb.create_sheet(title=batch_name)
        
        # Fixed headers
        fixed_headers = ["Sr No.", "Full Name", "Roll No.", "Branch", "Present", "Absent", "Attendance"]
        
        # Generate dates for 30 days (sample range)
        start_date = datetime(2025, 12, 15)
        dates = []
        date_objs = []
        for i in range(30):
            d = start_date + timedelta(days=i)
            if d.weekday() < 5:  # Skip weekends for sample
                dates.append(d.strftime("%d-%m-%Y"))
                date_objs.append(d)
        
        # Row 1: Headers with dates
        headers = fixed_headers + dates
        ws.append(headers)
        
        # Row 2: Session types
        session_headers = [""] * len(fixed_headers) + []
        for d in dates:
            session_headers.extend(["Morning Session", "Afternoon Session"])
        ws.append(session_headers[:len(dates)*2 + len(fixed_headers)])
        
        # Format header rows
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        for col in range(1, len(session_headers) + 1):
            cell = ws.cell(row=2, column=col)
            if cell.value:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Add sample student rows
        for i in range(1, 6):  # 5 sample students
            row = [i, f"Student {i}", f"22CS100{i}", "CSE", 25-i, i, f"{((25-i)*100)//30}%"]
            # Add attendance marks (sample)
            for _ in dates:
                row.extend(["P", "P"] if i % 3 != 0 else ["A", "P"])
            ws.append(row)
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 12
        
        for col in range(8, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 12
    
    wb.save("attendance_template.xlsx")
    print("✓ Sample Excel template created: attendance_template.xlsx")
    print("  Batches: Batch 1, Batch 2, Batch 3, Batch 4")
    print("  Each with 5 sample students and 30 days of data")

if __name__ == "__main__":
    create_template()
