from django.core.management.base import BaseCommand
from django.utils import timezone
from openpyxl import load_workbook

from students.models import Student
from lectures.models import Lecture, Batch
from attendance.models import AttendanceRecord


class Command(BaseCommand):
    help = "Import attendance from Excel sheets (batch-wise)"

    def add_arguments(self, parser):
        parser.add_argument("file_path", type=str)
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Validate data without writing to DB",
        )
        parser.add_argument(
            "--date",
            type=str,
            help="Lecture date (YYYY-MM-DD). Defaults to today",
        )

    def handle(self, *args, **options):
        file_path = options["file_path"]
        dry_run = options["dry_run"]
        date_str = options.get("date")

        lecture_date = (
            timezone.datetime.fromisoformat(date_str).date()
            if date_str
            else timezone.localdate()
        )

        wb = load_workbook(file_path)

        created = updated = skipped = 0

        for sheet_name in wb.sheetnames:
            try:
                batch = Batch.objects.get(name=sheet_name)
            except Batch.DoesNotExist:
                self.stderr.write(
                    self.style.WARNING(f"Skipping sheet {sheet_name} (no such batch)")
                )
                continue

            sheet = wb[sheet_name]
            #headers = [c.value for c in sheet[1]]
            headers = [str(c.value).strip().replace("\n", "").replace("\xa0", " ")for c in sheet[1]]


            required = ["Roll No.", "Morning", "Afternoon"]
            if not all(h in headers for h in required):
                self.stderr.write(
                    self.style.ERROR(f"Invalid headers in sheet {sheet_name}")
                )
                continue

            idx = {h: headers.index(h) for h in headers}

            for row in sheet.iter_rows(min_row=2, values_only=True):
                roll_no = row[idx["Roll No."]]
                morning = row[idx["Morning"]]
                afternoon = row[idx["Afternoon"]]

                if not roll_no:
                    skipped += 1
                    continue

                try:
                    student = Student.objects.get(roll_number=roll_no, batch=batch)
                except Student.DoesNotExist:
                    skipped += 1
                    self.stderr.write(
                        self.style.WARNING(f"Student not found: {roll_no}")
                    )
                    continue

                for session, value in [("MS", morning), ("AS", afternoon)]:
                    if value not in ("P", "A"):
                        skipped += 1
                        continue

                    try:
                        lecture = Lecture.objects.get(
                            batch=batch,
                            date=lecture_date,
                            lecture_type=session,
                        )
                    except Lecture.DoesNotExist:
                        skipped += 1
                        self.stderr.write(
                            self.style.WARNING(
                                f"No lecture for {batch.name} {lecture_date} {session}"
                            )
                        )
                        continue

                    if dry_run:
                        continue

                    obj, is_created = AttendanceRecord.objects.update_or_create(
                        student=student,
                        lecture=lecture,
                        defaults={
                            "status": value,
                        },
                    )

                    created += int(is_created)
                    updated += int(not is_created)

        if dry_run:
            self.stdout.write(
                self.style.SUCCESS(
                    "DRY RUN COMPLETE — no database changes were made"
                )
            )

        self.stdout.write(
            self.style.SUCCESS(
                f"Done → Created: {created}, Updated: {updated}, Skipped: {skipped}"
            )
        )
