from django.core.management.base import BaseCommand
from students.models import Student
from lectures.models import Batch, Slot
from openpyxl import load_workbook
from django.db import transaction


class Command(BaseCommand):
    help = "Import students from Excel file"

    def add_arguments(self, parser):
        parser.add_argument("file", type=str)
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Validate only, do not write to DB",
        )

    def handle(self, *args, **options):
        file_path = options["file"]
        dry_run = options["dry_run"]

        wb = load_workbook(file_path)

        required_headers = [
            "Full Name",
            "Roll No.",
            "Branch",
            "Email",
            "Contact",
            "Parent Email",
            "Parent Contact",
        ]

        created = 0
        errors = []

        slot_map = {
            "Batch 1": "Slot 1",
            "Batch 2": "Slot 1",
            "Batch 3": "Slot 2",
            "Batch 4": "Slot 3",
        }

        try:
            with transaction.atomic():
                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]

                    headers = [
                        str(c.value).strip().replace("\n", "").replace("\xa0", " ")
                        for c in sheet[1]
                    ]

                    if not all(h in headers for h in required_headers):
                        errors.append(
                            f"{sheet_name}: Invalid headers → {headers}"
                        )
                        continue

                    idx = {h: headers.index(h) for h in headers}

                    try:
                        batch = Batch.objects.get(name=sheet_name)
                        slot = Slot.objects.get(name=slot_map[sheet_name])
                    except Exception:
                        errors.append(f"{sheet_name}: Batch or Slot missing")
                        continue

                    for row_no, row in enumerate(sheet.iter_rows(min_row=2), start=2):
                        roll = str(row[idx["Roll No."]].value).strip()

                        if not roll:
                            errors.append(f"{sheet_name} Row {row_no}: Missing roll number")
                            continue

                        if Student.objects.filter(roll_number=roll).exists():
                            errors.append(
                                f"{sheet_name} Row {row_no}: Duplicate roll {roll}"
                            )
                            continue

                        Student(
                            full_name=row[idx["Full Name"]].value,
                            roll_number=roll,
                            branch=row[idx["Branch"]].value,
                            email=row[idx["Email"]].value,
                            contact_number=str(row[idx["Contact"]].value),
                            parent_email=row[idx["Parent Email"]].value,
                            parent_contact_number=str(
                                row[idx["Parent Contact"]].value
                            ),
                            batch=batch,
                            slot=slot,
                        ).full_clean()

                        if not dry_run:
                            Student.objects.create(
                                full_name=row[idx["Full Name"]].value,
                                roll_number=roll,
                                branch=row[idx["Branch"]].value,
                                email=row[idx["Email"]].value,
                                contact_number=str(row[idx["Contact"]].value),
                                parent_email=row[idx["Parent Email"]].value,
                                parent_contact_number=str(
                                    row[idx["Parent Contact"]].value
                                ),
                                batch=batch,
                                slot=slot,
                            )

                        created += 1

                if errors:
                    raise Exception("Validation failed")

                if dry_run:
                    raise transaction.TransactionManagementError("Dry run rollback")

        except Exception:
            pass

        if errors:
            self.stderr.write(self.style.ERROR("IMPORT FAILED"))
            for e in errors:
                self.stderr.write(f" - {e}")
            return

        if dry_run:
            self.stdout.write(
                self.style.WARNING(
                    f"DRY RUN OK → {created} students validated (no DB writes)"
                )
            )
        else:
            self.stdout.write(
                self.style.SUCCESS(f"IMPORT SUCCESS → {created} students added")
            )
