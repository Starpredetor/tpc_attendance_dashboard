from django.core.management.base import BaseCommand, CommandError
from django.utils import timezone
from datetime import datetime
from openpyxl import load_workbook
from lectures.models import Batch, Lecture
from students.models import Student
from attendance.models import AttendanceRecord
from django.contrib.auth import get_user_model
import re

User = get_user_model()


class Command(BaseCommand):
    help = "Import attendance from Excel file with format: dates as columns, students as rows"

    def add_arguments(self, parser):
        parser.add_argument("file_path", type=str, help="Path to the Excel file")
        parser.add_argument(
            "--clear",
            action="store_true",
            help="Clear existing attendance records before import",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Preview import without making changes to database",
        )

    def handle(self, *args, **options):
        file_path = options["file_path"]
        clear = options.get("clear", False)
        dry_run = options.get("dry_run", False)

        if dry_run:
            self.stdout.write(self.style.WARNING("DRY-RUN MODE: No changes will be made to the database"))

        try:
            wb = load_workbook(file_path)
        except Exception as e:
            raise CommandError(f"Failed to load Excel file: {e}")

        if clear and not dry_run:
            AttendanceRecord.objects.all().delete()
            Lecture.objects.all().delete()
            self.stdout.write(self.style.WARNING("Cleared all attendance and lecture records"))
        elif clear and dry_run:
            self.stdout.write(self.style.WARNING("[DRY-RUN] Would clear all attendance and lecture records"))

        # Get or create a system user for lectures
        system_user, _ = User.objects.get_or_create(
            email="admin@tpc.com",
            defaults={"full_name": "System Import", "role": "ADMIN"},
        )

        # Track statistics
        stats = {
            "batches_processed": 0,
            "lectures_created": 0,
            "attendance_records": 0,
            "students_not_found": 0,
            "batches_not_found": 0,
        }

        for sheet_name in wb.sheetnames:
            self.stdout.write(f"\nProcessing sheet: {sheet_name}")
            ws = wb[sheet_name]

            # Extract batch name from sheet (e.g., "Batch 1", "batch1", etc.)
            batch_name = sheet_name.strip()
            try:
                batch = Batch.objects.get(name=batch_name)
                stats["batches_processed"] += 1
            except Batch.DoesNotExist:
                self.stdout.write(
                    self.style.ERROR(f"Batch '{batch_name}' not found. Skipping sheet.")
                )
                stats["batches_not_found"] += 1
                continue

            # Parse header to extract dates and session types
            # Row 1: Sr No., Full Name, Roll No., Branch, Present, Absent, Attendance, then DATE columns
            # Row 2: blank, blank, blank, blank, blank, blank, blank, then SESSION TYPE (Morning Session / Afternoon Session / Holiday)
            # Use explicit cell access for reliability
            
            fixed_cols = 7  # Sr No., Full Name, Roll No., Branch, Present, Absent, Attendance (columns A-G)

            # Extract dates and sessions from columns H onwards (col_idx 8+)
            date_sessions = []
            col_idx = fixed_cols + 1  # Start from column H (1-based indexing)
            last_date_obj = None  # Track the last date for None columns (afternoon sessions)
            
            while col_idx <= ws.max_column:
                # Get date from row 1 and session type from row 2
                date_cell_value = ws.cell(row=1, column=col_idx).value
                session_cell_value = ws.cell(row=2, column=col_idx).value
                
                # Determine the date for this column
                if date_cell_value:
                    # New date found
                    date_obj = self._parse_date(date_cell_value)
                    last_date_obj = date_obj
                else:
                    # No date, use the last date (for afternoon sessions paired with morning)
                    date_obj = last_date_obj
                
                # Process if we have both a date and a session
                if date_obj and session_cell_value:
                    session_str = str(session_cell_value).strip().lower()
                    
                    # Check for holidays
                    date_str = str(date_cell_value if date_cell_value else "").strip().lower()
                    if "holiday" in date_str or "holiday" in session_str:
                        # Skip holidays
                        col_idx += 1
                        continue
                    elif "morning" in session_str:
                        session_type = "MS"
                    elif "afternoon" in session_str:
                        session_type = "AS"
                    else:
                        col_idx += 1
                        continue

                    # Add this date/session combo
                    date_sessions.append({
                        "date": date_obj,
                        "session": session_type,
                        "col": col_idx
                    })

                col_idx += 1

            self.stdout.write(f"Found {len(date_sessions)} date/session combinations")

            # Create lectures for each date/session
            lectures_by_date_session = {}
            for ds in date_sessions:
                key = (ds["date"], ds["session"])
                if key not in lectures_by_date_session:
                    if not dry_run:
                        lecture, created = Lecture.objects.get_or_create(
                            batch=batch,
                            date=ds["date"],
                            lecture_type=ds["session"],
                            defaults={
                                "title": f"{batch_name} - {ds['session']} - {ds['date'].strftime('%d-%b-%Y')}",
                                "created_by": system_user,
                            },
                        )
                        lectures_by_date_session[key] = lecture
                        if created:
                            self.stdout.write(
                                f"  [CREATE] Lecture: {ds['date'].strftime('%d-%b-%Y')} {ds['session']}"
                            )
                            stats["lectures_created"] += 1
                        else:
                            self.stdout.write(
                                f"  [FOUND] Lecture: {ds['date'].strftime('%d-%b-%Y')} {ds['session']}"
                            )
                    else:
                        # In dry-run, create a mock lecture object for tracking
                        mock_lecture = type('Lecture', (), {
                            'date': ds['date'],
                            'lecture_type': ds['session'],
                            'batch': batch
                        })()
                        lectures_by_date_session[key] = mock_lecture
                        self.stdout.write(
                            f"  [WOULD CREATE] Lecture: {ds['date'].strftime('%d-%b-%Y')} {ds['session']}"
                        )
                        stats["lectures_created"] += 1

            # Parse student rows and mark attendance
            # Roll number is in column C (index 2), starting from row 3
            row_num = 3  # Data starts at row 3
            while row_num <= ws.max_row:
                # Explicitly read from columns A-G (fixed columns)
                try:
                    roll_number_cell = ws.cell(row=row_num, column=3).value  # Column C
                    if not roll_number_cell:
                        row_num += 1
                        continue
                        
                    roll_number = str(roll_number_cell).strip()
                    if not roll_number:
                        row_num += 1
                        continue
                except Exception:
                    row_num += 1
                    continue

                try:
                    student = Student.objects.get(roll_number=roll_number, batch=batch)
                except Student.DoesNotExist:
                    self.stdout.write(
                        self.style.WARNING(f"  Student {roll_number} not found in batch {batch_name}")
                    )
                    stats["students_not_found"] += 1
                    row_num += 1
                    continue

                # Mark attendance for each date/session
                for i, ds in enumerate(date_sessions):
                    col_idx = ds["col"]  # Already 1-based (openpyxl indexing)
                    try:
                        status_cell = ws.cell(row=row_num, column=col_idx).value
                        if status_cell:
                            status_str = str(status_cell).strip().upper()
                            if status_str in ("P", "PRESENT"):
                                status = "P"
                            elif status_str in ("A", "ABSENT"):
                                status = "A"
                            elif "HOLIDAY" in status_str:
                                status = None  # Skip holidays
                            else:
                                status = None

                            if status:
                                key = (ds["date"], ds["session"])
                                lecture = lectures_by_date_session.get(key)
                                if lecture:
                                    if not dry_run:
                                        try:
                                            AttendanceRecord.objects.update_or_create(
                                                student=student,
                                                lecture=lecture,
                                                defaults={"status": status, "marked_by": system_user},
                                            )
                                            stats["attendance_records"] += 1
                                        except Exception as e:
                                            self.stdout.write(
                                                self.style.ERROR(
                                                    f"  ERROR creating attendance for {student.roll_number} "
                                                    f"on {ds['date']}: {e}"
                                                )
                                            )
                                    else:
                                        stats["attendance_records"] += 1
                                elif not dry_run:
                                    self.stdout.write(
                                        self.style.WARNING(
                                            f"  WARNING: No lecture found for {ds['date']} {ds['session']}"
                                        )
                                    )
                    except Exception as e:
                        self.stdout.write(
                            self.style.ERROR(
                                f"  ERROR processing attendance cell at row {row_num}, "
                                f"col {col_idx}: {e}"
                            )
                        )

                row_num += 1

            self.stdout.write(self.style.SUCCESS(f"Completed processing for {batch_name}"))

        # Print summary
        self.stdout.write("\n" + "="*60)
        self.stdout.write(self.style.SUCCESS("IMPORT SUMMARY"))
        self.stdout.write("="*60)
        self.stdout.write(f"Batches processed:        {stats['batches_processed']}")
        self.stdout.write(f"Batches not found:        {stats['batches_not_found']}")
        self.stdout.write(f"Lectures created/found:   {stats['lectures_created']}")
        self.stdout.write(f"Attendance records:       {stats['attendance_records']}")
        self.stdout.write(f"Students not found:       {stats['students_not_found']}")
        self.stdout.write("="*60)

        if dry_run:
            self.stdout.write(
                self.style.WARNING(
                    "\nDRY-RUN COMPLETED: No changes were made to the database.\n"
                    "Run without --dry-run to actually import the data."
                )
            )
        else:
            self.stdout.write(self.style.SUCCESS("\nAttendance import completed successfully!"))

    def _parse_date(self, date_str):
        """Attempt to parse a date string in multiple formats, or convert datetime objects."""
        # If already a datetime object, just extract the date
        if hasattr(date_str, 'date'):
            # It's a datetime.datetime object
            return date_str.date()
        
        # If already a date object, return as-is
        if hasattr(date_str, 'year') and hasattr(date_str, 'month') and hasattr(date_str, 'day'):
            if not hasattr(date_str, 'hour'):
                # It's a date object
                return date_str
        
        # Try parsing as string
        formats = [
            "%d-%m-%Y",
            "%d/%m/%Y",
            "%d-%b-%Y",
            "%d %b %Y",
            "%Y-%m-%d",
            "%m-%d-%Y",
        ]
        for fmt in formats:
            try:
                return datetime.strptime(str(date_str), fmt).date()
            except ValueError:
                continue
        return None
